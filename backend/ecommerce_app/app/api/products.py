from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from ..schemas.product import ProductCreate, ProductOut, ProductUpdate
from ..crud.product import list_products, get_product, get_product_by_sku, create_product, update_product, delete_product, search_products
from ..crud.category import collect_descendant_ids
from ..database import get_db
from ..dependencies import require_upload_api_key
from pydantic import BaseModel
import csv, io
from typing import Any

router = APIRouter()

@router.get('/', response_model=list[ProductOut])
def list_all(
    skip: int = 0,
    limit: int = 20,
    q: str | None = None,
    category: str | None = None,  # legacy string
    category_id: int | None = None,
    category_path: str | None = None,
    include_descendants: bool = True,
    db: Session = Depends(get_db)
):
    # If category_path provided, ensure it exists and use its id
    if category_path:
        from ..crud.category import ensure_hierarchy
        cat = ensure_hierarchy(db, category_path)
        category_id = cat.id
    # Expand category_id to include descendants if requested
    cat_ids = None
    if category_id is not None and include_descendants:
        try:
            cat_ids = collect_descendant_ids(db, category_id, include_self=True)
        except Exception:
            cat_ids = [category_id]
    if q:
        return search_products(db, q, skip, limit, category, category_id if not cat_ids else None, cat_ids)
    return list_products(db, skip, limit, category, category_id if not cat_ids else None, cat_ids)

@router.post('/', response_model=ProductOut, dependencies=[Depends(require_upload_api_key)])
def create(product_in: ProductCreate, db: Session = Depends(get_db)):
    return create_product(db, product_in)

@router.get('/{product_id}', response_model=ProductOut)
def retrieve(product_id: int, db: Session = Depends(get_db)):
    product = get_product(db, product_id)
    if not product:
        raise HTTPException(status_code=404, detail='Ürün bulunamadı')
    return product


@router.get('/by-sku/{sku}', response_model=ProductOut)
def retrieve_by_sku(sku: str, db: Session = Depends(get_db)):
    product = get_product_by_sku(db, sku)
    if not product:
        raise HTTPException(status_code=404, detail='Ürün bulunamadı')
    return product



@router.put('/{product_id}', response_model=ProductOut)
def update(product_id: int, product_in: ProductUpdate, db: Session = Depends(get_db)):
    product = get_product(db, product_id)
    if not product:
        raise HTTPException(status_code=404, detail='Ürün bulunamadı')
    return update_product(db, product, product_in)

@router.delete('/{product_id}')
def remove(product_id: int, db: Session = Depends(get_db)):
    product = get_product(db, product_id)
    if not product:
        raise HTTPException(status_code=404, detail='Ürün bulunamadı')
    delete_product(db, product)
    return {'detail': 'Silindi'}


class AssignCategoryPayload(BaseModel):
    sku: str
    category_path: str


@router.post('/assign-category')
def assign_category(payload: AssignCategoryPayload, db: Session = Depends(get_db)):
    from sqlalchemy import select
    from ..models.product import Product
    from ..crud.category import ensure_hierarchy
    prod = db.scalar(select(Product).where(Product.sku == payload.sku))
    if not prod:
        raise HTTPException(status_code=404, detail='Ürün bulunamadı (sku)')
    cat = ensure_hierarchy(db, payload.category_path)
    prod.category_id = cat.id
    db.add(prod)
    db.commit()
    db.refresh(prod)
    return {'detail': 'Kategori atandı', 'product_id': prod.id, 'category_id': cat.id}


class BulkAssignItem(BaseModel):
    sku: str
    category_path: str


@router.post('/assign-category-bulk')
def assign_category_bulk(items: list[BulkAssignItem], db: Session = Depends(get_db)):
    from sqlalchemy import select
    from ..models.product import Product
    from ..crud.category import ensure_hierarchy
    results = []
    for it in items:
        prod = db.scalar(select(Product).where(Product.sku == it.sku))
        if not prod:
            results.append({'sku': it.sku, 'status': 'not_found'})
            continue
        cat = ensure_hierarchy(db, it.category_path)
        prod.category_id = cat.id
        db.add(prod)
        results.append({'sku': it.sku, 'category_id': cat.id, 'status': 'ok'})
    db.commit()
    return {'updated': [r for r in results if r.get('status') == 'ok'], 'not_found': [r for r in results if r.get('status') == 'not_found']}


def _parse_excel(file_bytes: bytes) -> list[dict[str, Any]]:
    # Lazy import openpyxl to avoid hard dependency at import time
    import openpyxl  # type: ignore
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    # First row headers
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else '')
    rows: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        obj = {}
        for idx, val in enumerate(r):
            key = headers[idx] if idx < len(headers) else f'col_{idx}'
            obj[key] = val
        rows.append(obj)
    return rows


def _parse_csv(file_bytes: bytes, encoding: str = 'utf-8') -> list[dict[str, Any]]:
    text = file_bytes.decode(encoding, errors='ignore')
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _norm_key(s: str | None) -> str:
    return (s or '').strip().lower()


def _norm_url(v: Any) -> str | None:
    """Ensure image-like values are absolute URLs with scheme.
    - Keep http://, https://, data: as-is
    - If starts with //, prefix https:
    - Otherwise, prefix https:// and strip leading slashes
    """
    if v in (None, ''):
        return None
    s = str(v).strip()
    if not s:
        return None
    low = s.lower()
    if low.startswith('http://') or low.startswith('https://') or low.startswith('data:'):
        return s
    if s.startswith('//'):
        return 'https:' + s
    # Default: assume domain/path and prefix https
    return 'https://' + s.lstrip('/')


@router.post('/import-excel', dependencies=[Depends(require_upload_api_key)])
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    ext = (file.filename or '').split('.')[-1].lower()
    if ext in ('xlsx', 'xlsm', 'xltx', 'xltm'):
        rows = _parse_excel(content)
    elif ext in ('csv',):
        rows = _parse_csv(content)
    else:
        raise HTTPException(status_code=400, detail='Sadece .xlsx veya .csv dosyaları desteklenir')

    # Accept the following headers (case-insensitive):
    # parent_category, child_category, subchild_category,
    # sku, title, product_description, feature1..feature8, brand, main_img, img1..img4, meta_title, meta_description, schema_description
    updated = 0
    created = 0
    errors: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        try:
            keymap = { _norm_key(k): k for k in row.keys() }
            def get_exact(cand: str):
                k = keymap.get(_norm_key(cand))
                if k and row.get(k) not in (None, ''):
                    return row[k]
                return None

            # Build category_path from exact parts
            parent = get_exact('parent_category')
            child = get_exact('child_category')
            sub = get_exact('subchild_category')
            parts = [p for p in [parent, child, sub] if p not in (None, '')]
            category_path = ' > '.join([str(p) for p in parts]) if parts else None

            sku_val = get_exact('sku')
            # Use only 'title' header; legacy 'name' is no longer supported. If missing, fallback to SKU below.
            title_val = get_exact('title')
            # 'product_description' (fallback to 'description' if present)
            description = get_exact('product_description') or get_exact('description')
            brand = get_exact('brand')
            # images (english-only)
            main_img = _norm_url(get_exact('main_img'))
            img1 = _norm_url(get_exact('img1'))
            img2 = _norm_url(get_exact('img2'))
            img3 = _norm_url(get_exact('img3'))
            img4 = _norm_url(get_exact('img4'))
            # features (feature1..feature8)
            features = {}
            for i in range(1, 9):
                val = get_exact(f'feature{i}')
                if val not in (None, ''):
                    features[f'feature{i}'] = str(val)
            # meta
            meta_title = get_exact('meta_title')
            meta_description = get_exact('meta_description')
            schema_description = get_exact('schema_description')

            # Build payload
            payload: dict[str, Any] = {}
            if title_val: payload['title'] = str(title_val)
            if description: payload['description'] = str(description)
            if sku_val: payload['sku'] = str(sku_val)
            if category_path: payload['category_path'] = str(category_path)
            if brand: payload['brand'] = str(brand)
            if main_img: payload['main_img'] = str(main_img)
            if img1: payload['img1'] = str(img1)
            if img2: payload['img2'] = str(img2)
            if img3: payload['img3'] = str(img3)
            if img4: payload['img4'] = str(img4)
            payload.update(features)
            if meta_title or meta_description or schema_description:
                payload['seo'] = {}
                if meta_title:
                    payload['seo']['meta_title'] = str(meta_title)
                if meta_description:
                    payload['seo']['meta_description'] = str(meta_description)
                if schema_description:
                    payload['seo']['schema_description'] = str(schema_description)

            # Auto-fill title from SKU if missing; require minimum: sku
            if not payload.get('sku'):
                raise ValueError('sku eksik')
            if not payload.get('title'):
                payload['title'] = payload['sku']

            # Upsert by SKU
            from sqlalchemy import select
            from ..models.product import Product
            existing = db.scalar(select(Product).where(Product.sku == str(sku_val).strip().upper()))
            if existing:
                # update via CRUD (handles nested seo and category_path)
                payload_no_code = {k: v for k, v in payload.items() if k != 'sku'}
                upd = ProductUpdate(**payload_no_code)
                update_product(db, existing, upd)
                updated += 1
            else:
                # create
                crt = ProductCreate(**payload)
                prod = create_product(db, crt)
                created += 1

            # content logic removed
        except Exception as e:
            errors.append({'row': idx, 'error': str(e)})
    db.commit()
    return {'created': created, 'updated': updated, 'errors': errors}
